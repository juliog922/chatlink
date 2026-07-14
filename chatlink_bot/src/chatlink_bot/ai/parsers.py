# src/chatlink_bot/ai/parsers.py
import asyncio
import re
import logging
import os
import base64
import tempfile
from io import BytesIO
from typing import Optional

from PIL import Image, ImageEnhance, ImageFilter, ExifTags

from .cima_client import CimaClient, Message

logger = logging.getLogger("Parsers")

CIMA_URL = os.getenv("CIMA_URL", "http://cima:8000")
# One GGUF model does text, vision AND audio — no separate VLM/ASR models.
CIMA_MODEL = os.getenv("CIMA_MODEL", "unsloth/gemma-4-E2B-it-GGUF:Q8_0")

# ---------------------------------------------------------------------------
# Image preprocessing tunables (override via env for quick experiments)
# ---------------------------------------------------------------------------
# Minimum dimension before upscaling kicks in.  WhatsApp thumbnails and
# heavily-compressed photos often land at 800-1200px on the longest side;
# Qwen2.5-VL performs noticeably better above ~1500px.
IMG_MIN_LONG_EDGE = int(os.getenv("IMG_MIN_LONG_EDGE", "1500"))

# Upper pixel-count cap BEFORE sending to the engine.  Keeps base64 payload
# and server-side tensor allocation sane.
IMG_MAX_PIXELS = int(os.getenv("IMG_MAX_PIXELS", str(30_000_000)))

# Hard cap on the longest edge. cima's decoder enforces dimension limits and
# the vision tower budget is finite on a 6 GiB card; 2048px keeps OCR quality
# while bounding VRAM and payload size.
IMG_MAX_LONG_EDGE = int(os.getenv("IMG_MAX_LONG_EDGE", "2048"))

# Sharpening strength applied via UnsharpMask to counteract JPEG blur.
# 0 = disabled.  Good range for WhatsApp photos: 1.0-2.0
IMG_SHARPEN_RADIUS = float(os.getenv("IMG_SHARPEN_RADIUS", "2"))
IMG_SHARPEN_PERCENT = int(os.getenv("IMG_SHARPEN_PERCENT", "150"))
IMG_SHARPEN_THRESHOLD = int(os.getenv("IMG_SHARPEN_THRESHOLD", "3"))

# Mild contrast boost (1.0 = no change, 1.1-1.3 helps washed-out photos)
IMG_CONTRAST_FACTOR = float(os.getenv("IMG_CONTRAST_FACTOR", "1.15"))

# Mild sharpness boost via Pillow's Sharpness enhancer (stacks with USM)
IMG_SHARPNESS_FACTOR = float(os.getenv("IMG_SHARPNESS_FACTOR", "1.3"))

_client: Optional[CimaClient] = None


def _get_client() -> CimaClient:
    global _client
    if _client is None:
        _client = CimaClient(CIMA_URL)
    return _client


# ---------------------------------------------------------------------------
# Image preprocessing pipeline for WhatsApp-compressed photos
# ---------------------------------------------------------------------------

def _fix_exif_orientation(img: Image.Image) -> Image.Image:
    """Apply EXIF orientation transforms so text isn't sideways."""
    try:
        exif = img.getexif()
        orientation_key = next(
            (tag_id for tag_id, name in ExifTags.TAGS.items() if name == "Orientation"),
            None,
        )
        if orientation_key and orientation_key in exif:
            _T = Image.Transpose
            transforms = {
                2: [_T.FLIP_LEFT_RIGHT],
                3: [_T.ROTATE_180],
                4: [_T.FLIP_TOP_BOTTOM],
                5: [_T.FLIP_LEFT_RIGHT, _T.ROTATE_90],
                6: [_T.ROTATE_270],
                7: [_T.FLIP_LEFT_RIGHT, _T.ROTATE_270],
                8: [_T.ROTATE_90],
            }
            for op in transforms.get(exif[orientation_key], []):
                img = img.transpose(op)
    except Exception:
        pass
    return img


def _ensure_rgb(img: Image.Image) -> Image.Image:
    """Convert any mode (RGBA, P, LA, L, CMYK …) to clean RGB."""
    if img.mode == "RGB":
        return img
    if img.mode in ("RGBA", "LA", "PA"):
        background = Image.new("RGB", img.size, (255, 255, 255))
        # Use alpha channel as mask if present
        if "A" in img.mode:
            rgba = img.convert("RGBA")
            background.paste(rgba, mask=rgba.split()[3])
        else:
            background.paste(img)
        return background
    if img.mode == "P":
        return _ensure_rgb(img.convert("RGBA"))
    # L, CMYK, etc.
    return img.convert("RGB")


def _preprocess_image(raw_bytes: bytes) -> bytes:
    """
    Full preprocessing pipeline for WhatsApp images before VLM OCR:
    
    1. Decode & validate
    2. Fix EXIF orientation (WhatsApp/phone rotation)
    3. Convert to RGB
    4. Upscale small images (Qwen2.5-VL needs sufficient resolution)
    5. Apply UnsharpMask (counteract JPEG compression blur)
    6. Mild contrast & sharpness boost
    7. Encode as lossless PNG (avoid re-compression artifacts)
    
    Returns PNG bytes ready for base64 encoding.
    """
    img = Image.open(BytesIO(raw_bytes))
    img.load()  # Force full decode — catches truncated files early

    # --- EXIF orientation (critical for phone photos) ---
    img = _fix_exif_orientation(img)

    # --- RGB conversion ---
    img = _ensure_rgb(img)

    w, h = img.size

    # --- Downscale if above pixel cap (safety valve) ---
    if w * h > IMG_MAX_PIXELS:
        scale = (IMG_MAX_PIXELS / (w * h)) ** 0.5
        new_w, new_h = int(w * scale), int(h * scale)
        img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
        w, h = new_w, new_h
        logger.debug("Downscaled to %dx%d (pixel cap)", w, h)

    # --- Upscale small images ---
    # WhatsApp often delivers ~800-1200px.  Qwen2.5-VL's dynamic resolution
    # slicing performs better when text characters span more pixels.
    long_edge = max(w, h)
    if long_edge < IMG_MIN_LONG_EDGE and long_edge > 0:
        scale = IMG_MIN_LONG_EDGE / long_edge
        new_w, new_h = int(w * scale), int(h * scale)
        img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
        logger.debug("Upscaled from %dx%d to %dx%d", w, h, new_w, new_h)

    # --- Long-edge cap (applies AFTER upscale so both paths are bounded) ---
    w, h = img.size
    long_edge = max(w, h)
    if long_edge > IMG_MAX_LONG_EDGE:
        scale = IMG_MAX_LONG_EDGE / long_edge
        img = img.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
        logger.debug("Capped long edge to %d", IMG_MAX_LONG_EDGE)

    # --- UnsharpMask (counteracts JPEG blur, the #1 OCR killer) ---
    if IMG_SHARPEN_RADIUS > 0:
        img = img.filter(
            ImageFilter.UnsharpMask(
                radius=IMG_SHARPEN_RADIUS,
                percent=IMG_SHARPEN_PERCENT,
                threshold=IMG_SHARPEN_THRESHOLD,
            )
        )

    # --- Contrast boost (helps washed-out/low-light photos) ---
    if IMG_CONTRAST_FACTOR != 1.0:
        img = ImageEnhance.Contrast(img).enhance(IMG_CONTRAST_FACTOR)

    # --- Sharpness boost (stacks with USM for fine text edges) ---
    if IMG_SHARPNESS_FACTOR != 1.0:
        img = ImageEnhance.Sharpness(img).enhance(IMG_SHARPNESS_FACTOR)

    # --- Encode as lossless PNG ---
    buf = BytesIO()
    img.save(buf, format="PNG", optimize=False)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Qwen 2.5-VL optimized OCR prompt
# ---------------------------------------------------------------------------

_OCR_SYSTEM_PROMPT = (
    "Eres el lector de imágenes de un asistente que toma PEDIDOS DE COSMÉTICA "
    "por WhatsApp. Las imágenes de los clientes casi siempre contienen pedidos: "
    "listas manuscritas, etiquetas de productos, fotos de envases o estanterías, "
    "capturas de pedidos anteriores.\n"
    "TAREA, en este orden:\n"
    "1) Transcribe fielmente TODO el texto visible, preservando estructura "
    "(saltos de línea, listas, tablas). No corrijas ortografía; mantén el idioma "
    "original; marca lo dudoso como [ilegible].\n"
    "2) Si detectas productos, termina con una sección:\n"
    "PRODUCTOS DETECTADOS:\n"
    "- <cantidad si se ve> <producto> <marca si se ve> <código/referencia si se ve> "
    "<color/variante si se ve>\n"
    "(una línea por producto; omite la sección si no hay productos).\n"
    "3) Si la imagen NO contiene texto: describe en 1-2 frases centrándote en los "
    "productos cosméticos visibles (tipo de producto, marca, color del envase o "
    "etiqueta), no en el fondo ni la escena."
)

_OCR_USER_PROMPT = (
    "Extrae el texto de esta imagen y, si contiene productos de cosmética, "
    "añade al final la sección PRODUCTOS DETECTADOS con cada producto, cantidad, "
    "marca y código visibles."
)


def extract_text_from_image_bytes(image_bytes: bytes) -> str:
    """
    Full pipeline: preprocess WhatsApp-compressed image → send to Qwen2.5-VL.
    
    Improvements over raw passthrough:
    - EXIF orientation fix (sideways phone photos)
    - Upscale small images for better VLM resolution slicing
    - UnsharpMask + contrast to counteract JPEG compression artifacts
    - Lossless PNG encoding (no re-compression)
    - Structured system + user prompt tuned for Qwen2.5-VL OCR
    - temperature=0 for deterministic, faithful extraction
    """
    if not image_bytes:
        return ""

    client = _get_client()
    try:
        # --- Preprocess: enhance the WhatsApp-compressed image ---
        # This ALSO normalizes the container: WhatsApp can deliver WebP
        # (stickers) or progressive JPEG, which the engine's baseline-only
        # decoder rejects. Pillow re-decodes anything and we re-encode to
        # lossless PNG, so what reaches cima is always decodable.
        try:
            clean_png = _preprocess_image(image_bytes)
        except Exception as e:
            # Preprocessing failed. Only fall back to raw bytes if they are
            # already a format cima decodes (PNG or JPEG). Sending unknown
            # containers (WebP, HEIC, truncated files) just errors server-side.
            magic_ok = image_bytes[:8] == b"\x89PNG\r\n\x1a\n" or image_bytes[:3] == b"\xff\xd8\xff"
            if magic_ok:
                logger.warning(f"Image preprocessing failed, sending raw PNG/JPEG: {e}")
                clean_png = image_bytes
            else:
                logger.error(f"Image undecodable (not PNG/JPEG after Pillow failure): {e}")
                return ""

        b64_img = base64.b64encode(clean_png).decode("utf-8")

        resp = client.chat(
            model=CIMA_MODEL,
            messages=[
                Message(role="system", content=_OCR_SYSTEM_PROMPT),
                Message(
                    role="user",
                    content=_OCR_USER_PROMPT,
                    images=[b64_img],
                ),
            ],
            options={
                "temperature": 0.0,   # Deterministic — faithful to the image
                "top_p": 1.0,         # No nucleus sampling for OCR
                "repeat_penalty": 1.0, # Don't penalize repeated text (tables, lists)
                "num_predict": 2048,   # Allow long extractions (invoices, documents)
            },
        )
        return (resp.content or "").strip()
    except Exception as e:
        logger.error(f"Vision OCR failed: {e}")
        return ""


_ASR_SYSTEM_PROMPT = (
    "Eres un motor de transcripción de audio de alta precisión. Tu ÚNICA tarea "
    "es transcribir fielmente el habla del audio a texto, en el idioma original "
    "(normalmente español). No traduzcas, no resumas, no respondas al hablante, "
    "no añadas comentarios: devuelve SOLO la transcripción literal. "
    "Si el audio es inentendible o no contiene habla, responde exactamente: ###SIN_HABLA###"
)

_ASR_USER_PROMPT = "Transcribe este audio literalmente, en su idioma original."

# ---------------------------------------------------------------------------
# Audio normalization — WhatsApp voice notes are Opus-in-OGG; emails carry
# mp3/m4a/etc. cima decodes ONLY PCM WAV, so anything else must be transcoded
# BEFORE upload or the request fails. ffmpeg does the conversion; without it
# non-WAV audio is rejected here with a clear log instead of collapsing the
# engine with an undecodable payload.
# ---------------------------------------------------------------------------

_AUDIO_TARGET_RATE = int(os.getenv("AUDIO_TARGET_RATE", "16000"))
_AUDIO_MAX_SECONDS = int(os.getenv("AUDIO_MAX_SECONDS", "300"))  # cap runaway clips


def _sniff_audio_kind(b: bytes) -> str:
    """Identify the container by magic bytes (never trust the filename)."""
    if len(b) < 12:
        return "unknown"
    if b[0:4] == b"RIFF" and b[8:12] == b"WAVE":
        return "wav"
    if b[0:4] == b"OggS":
        return "ogg"          # WhatsApp voice notes (Opus)
    if b[0:3] == b"ID3" or (b[0] == 0xFF and (b[1] & 0xE0) == 0xE0):
        return "mp3"
    if b[0:4] == b"fLaC":
        return "flac"
    if b[4:8] == b"ftyp":
        return "m4a"          # MP4/AAC family
    return "unknown"


def _wav_is_pcm16_mono_16k(b: bytes) -> bool:
    """True only for WAVs cima ingests optimally (PCM, mono-ish is fine)."""
    try:
        # fmt chunk: bytes 20-21 codec (1 = PCM), 22-23 channels, 24-27 rate
        codec = int.from_bytes(b[20:22], "little")
        return codec == 1
    except Exception:
        return False


def _wav_duration_seconds(wav: bytes) -> float:
    """Duration of a 16 kHz mono PCM16 WAV (the only shape we send)."""
    import struct
    if len(wav) < 44 or wav[:4] != b"RIFF":
        return 0.0
    pos = 12
    while pos + 8 <= len(wav):
        cid = bytes(wav[pos:pos + 4])
        csz = struct.unpack_from("<I", wav, pos + 4)[0]
        if cid == b"data":
            return (csz / 2.0) / float(_AUDIO_TARGET_RATE)
        pos += 8 + csz + (csz & 1)
    return 0.0


def _asr_quality(text: str, duration_s: float) -> str:
    """
    'ok' | 'suspect' | 'junk'. A small multimodal model hallucinates on hard
    audio and the agent must not treat that output as the client's words:
    - explicit no-speech token, empty, or non-linguistic output -> junk
    - heavy word repetition (classic ASR hallucination) -> junk
    - far fewer words than the duration implies (Spanish ~2-3 w/s) -> suspect
    """
    t = (text or "").strip()
    if not t or "###SIN_HABLA###" in t:
        return "junk"
    words = [w for w in re.split(r"\s+", t) if w]
    alpha = sum(ch.isalpha() for ch in t)
    if alpha < max(4, len(t) * 0.4):
        return "junk"
    if len(words) >= 8 and len(set(w.lower() for w in words)) / len(words) < 0.3:
        return "junk"                      # "gracias gracias gracias ..."
    if duration_s >= 4.0 and len(words) < duration_s * 0.5:
        return "suspect"                   # 10 s of speech, 3 words out
    return "ok"


def _finalize_wav_header(wav: bytes) -> bytes:
    """
    ffmpeg writing WAV to a PIPE cannot seek back to patch the RIFF and
    'data' chunk sizes, so it leaves 0xFFFFFFFF placeholders. Strict decoders
    (cima) then reject the file: "wav: chunk 'data' truncated" — observed
    live on every WhatsApp voice note. Rewrite both sizes from the actual
    byte counts; already-correct WAVs pass through byte-identical.
    """
    import struct
    if len(wav) < 44 or wav[:4] != b"RIFF" or wav[8:12] != b"WAVE":
        return wav
    total = len(wav)
    out = bytearray(wav)
    struct.pack_into("<I", out, 4, total - 8)          # RIFF chunk size
    pos = 12
    while pos + 8 <= total:
        chunk_id = bytes(out[pos:pos + 4])
        chunk_size = struct.unpack_from("<I", out, pos + 4)[0]
        if chunk_id == b"data":
            struct.pack_into("<I", out, pos + 4, total - (pos + 8))
            break
        pos += 8 + chunk_size + (chunk_size & 1)       # chunks are word-aligned
    return bytes(out)


def _ffmpeg_transcode_to_wav(raw: bytes, kind: str) -> bytes:
    """
    Transcode any container to 16 kHz mono PCM16 WAV via ffmpeg (stdin→stdout,
    no temp files). Raises with a clear message if ffmpeg is unavailable.
    """
    import subprocess

    cmd = [
        "ffmpeg", "-hide_banner", "-loglevel", "error",
        "-i", "pipe:0",
        "-t", str(_AUDIO_MAX_SECONDS),
        "-ac", "1", "-ar", str(_AUDIO_TARGET_RATE),
        "-c:a", "pcm_s16le", "-f", "wav", "pipe:1",
    ]
    try:
        proc = subprocess.run(
            cmd, input=raw, capture_output=True, timeout=120, check=False
        )
    except FileNotFoundError:
        raise RuntimeError(
            "ffmpeg is not installed in the bot image but is required to "
            f"transcode {kind} audio (WhatsApp sends Opus/OGG). "
            "Add `ffmpeg` to the chatlink_bot Dockerfile apt-get install."
        )
    except subprocess.TimeoutExpired:
        raise RuntimeError("ffmpeg timed out transcoding audio")
    if proc.returncode != 0 or not proc.stdout:
        raise RuntimeError(f"ffmpeg failed on {kind} audio: {proc.stderr.decode(errors='replace')[:300]}")
    return _finalize_wav_header(proc.stdout)


def _normalize_audio_for_cima(raw: bytes) -> bytes:
    """
    Ensure the payload is a PCM WAV cima can decode. WhatsApp compression
    (Opus/OGG) and email formats (mp3/m4a/flac) are transcoded; PCM WAV
    passes through untouched; non-PCM WAV is re-encoded.
    """
    kind = _sniff_audio_kind(raw)
    if kind == "wav" and _wav_is_pcm16_mono_16k(raw):
        return _finalize_wav_header(raw)
    logger.info(f"[AUDIO_NORM] Transcoding {kind} -> {_AUDIO_TARGET_RATE}Hz mono PCM WAV")
    return _ffmpeg_transcode_to_wav(raw, kind)


def transcribe_audio_bytes(audio_bytes: bytes, filename: str = "audio.wav") -> str:
    """
    Transcribe audio via cima's /api/generate (the only endpoint that accepts audio per the API spec).
    The clip is normalized to PCM WAV first — cima rejects Opus/OGG/MP3, which
    is what WhatsApp and email actually deliver.
    """
    if not audio_bytes:
        return ""

    client = _get_client()
    try:
        wav = _normalize_audio_for_cima(audio_bytes)
    except Exception as e:
        logger.error(f"Audio normalization failed (clip not transcribable): {e}")
        return ""

    try:
        b64_audio = base64.b64encode(wav).decode("utf-8")
        # Per cima's OpenAPI spec, `audio` exists ONLY on /api/generate
        # (top-level array); the chat Message schema has no audio field, so
        # attaching clips there made the server drop them silently and gemma
        # "transcribed" from imagination — observed live as fluent Spanish
        # unrelated to the voice note. GenerateRequest has no system field
        # either, so the instructions travel inline in the prompt.
        resp = client.generate(
            CIMA_MODEL,
            prompt=f"{_ASR_SYSTEM_PROMPT}\n\n{_ASR_USER_PROMPT}",
            audio=[b64_audio],
            options={
                "temperature": 0.0,
                "top_p": 1.0,
                "repeat_penalty": 1.0,
                "num_predict": 1024,
            },
        )
        text = (resp.content or "").strip()
        quality = _asr_quality(text, _wav_duration_seconds(wav))
        if quality == "junk":
            logger.warning(f"[ASR] Transcription judged JUNK "
                           f"({_wav_duration_seconds(wav):.1f}s audio): {text[:80]!r}")
            return ""
        if quality == "suspect":
            logger.warning(f"[ASR] Transcription judged SUSPECT "
                           f"({_wav_duration_seconds(wav):.1f}s audio): {text[:80]!r}")
            return f"(transcripción dudosa, puede tener errores) {text}"
        logger.info(f"[ASR] ok ({_wav_duration_seconds(wav):.1f}s): {text[:120]!r}")
        return text
    except Exception as e:
        logger.error(f"Audio transcription failed: {e}")
        return ""


def extract_text_from_document_bytes(data: bytes, filename: str) -> str:
    """
    Best-effort doc extraction for: txt/csv/json/md, pdf, docx, xlsx.
    """
    if not data:
        return ""

    fn = (filename or "").lower().strip()

    # text-like
    if fn.endswith((".txt", ".csv", ".md", ".json")):
        for enc in ("utf-8", "latin-1"):
            try:
                return data.decode(enc, errors="strict").strip()
            except Exception:
                continue
        return data.decode("utf-8", errors="replace").strip()

    # pdf
    if fn.endswith(".pdf"):
        try:
            from pypdf import PdfReader  # type: ignore

            reader = PdfReader(BytesIO(data))
            out = []
            for page in reader.pages:
                out.append(page.extract_text() or "")
            return "\n".join([x.strip() for x in out if x.strip()]).strip()
        except Exception as e:
            logger.warning(f"PDF extract failed: {e}")
            return ""

    # docx
    if fn.endswith(".docx"):
        try:
            from docx import Document  # type: ignore

            with tempfile.NamedTemporaryFile(suffix=".docx", delete=True) as f:
                f.write(data)
                f.flush()
                doc = Document(f.name)
            paras = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
            return "\n".join(paras).strip()
        except Exception as e:
            logger.warning(f"DOCX extract failed: {e}")
            return ""

    # xlsx/xlsm/xltx...
    if fn.endswith((".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")):
        try:
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
            out: list[str] = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
                    if vals:
                        out.append(" ".join(vals))
            return "\n".join(out).strip()
        except Exception as e:
            logger.warning(f"XLSX extract failed: {e}")
            return ""

    return ""


async def extract_text_from_image_bytes_async(image_bytes: bytes) -> str:
    return await asyncio.to_thread(extract_text_from_image_bytes, image_bytes)


async def transcribe_audio_bytes_async(audio_bytes: bytes, filename: str) -> str:
    return await asyncio.to_thread(transcribe_audio_bytes, audio_bytes, filename)


async def extract_text_from_document_bytes_async(data: bytes, filename: str) -> str:
    return await asyncio.to_thread(extract_text_from_document_bytes, data, filename)